"""
在线商品审计工具
使用AI审核在线商品信息（title, description, category, keyword）
没有原始URL，根据title和description判断商品合理性
"""

import csv
import json
import os
import sys
import traceback
from typing import Dict, List, Optional
import dashscope
from dashscope import Generation
from dotenv import load_dotenv
import pandas as pd
from datetime import datetime

# 加载环境变量（从项目根目录加载）
# 获取项目根目录（ai rating目录）
current_dir = os.path.dirname(os.path.abspath(__file__))
root_dir = os.path.dirname(current_dir)  # 上一级目录就是ai rating
env_path = os.path.join(root_dir, '.env')
load_dotenv(env_path)


class ProductAuditorOnline:
    """在线商品审计员（无URL版本）"""
    
    def __init__(self, api_key: str = None, model: str = "qwen-plus"):
        """初始化审计员
        
        Args:
            api_key: DashScope API Key
            model: 使用的模型名称，默认为 qwen-plus
        """
        self.api_key = api_key or os.getenv("QWEN_API_KEY") or os.getenv("DASHSCOPE_API_KEY")
        if not self.api_key:
            # 获取env_path用于错误提示
            current_dir = os.path.dirname(os.path.abspath(__file__))
            root_dir = os.path.dirname(current_dir)
            env_path = os.path.join(root_dir, '.env')
            raise ValueError(f"API Key未设置，请通过参数传入或设置环境变量QWEN_API_KEY或DASHSCOPE_API_KEY。环境变量文件位置: {env_path}")
        dashscope.api_key = self.api_key
        self.model = model
    
    def audit_product(self, offer_id: str, title: str, description: str, 
                     category_id: str, category_name: str, keywords: str) -> Dict:
        """审核单个商品
        
        Args:
            offer_id: 商品ID
            title: 商品标题
            description: 商品描述
            category_id: 分类ID
            category_name: 分类名称
            keywords: 关键词（JSON数组字符串）
            
        Returns:
            包含各项审核结果的字典
        """
        
        # 解析keywords
        keyword_text = ""
        keyword_list = []
        if keywords:
            try:
                if keywords.startswith('['):
                    keyword_list = json.loads(keywords)
                    if isinstance(keyword_list, list):
                        keyword_text = ', '.join([str(k) for k in keyword_list[:10]])  # 最多显示10个
                else:
                    keyword_text = keywords
            except:
                keyword_text = keywords
        
        # 构建审核prompt
        prompt = f"""You are an AI product auditor. Please review the following product information from an online platform.

**Product Information to Review:**

1. **Title:**
{title if title else "N/A"}

2. **Description:**
{description if description else "N/A"}

3. **Category:**
ID: {category_id if category_id else "N/A"}
Name: {category_name if category_name else "N/A"}

4. **Keywords:**
{keyword_text if keyword_text else "N/A"}

**Review Criteria:**

Since there is no source URL, you need to evaluate the product based on the title and description only.

**Product Validity (商品是否合理):**
This evaluation focuses on two key aspects:

1. **Non-Spam Content (非垃圾内容):**
   - Is the content meaningful and relevant?
   - Does it describe a real product?
   - Is it free of spam, gibberish, or placeholder text?
   - If content is spam, gibberish, or meaningless → mark as NEEDS_MANUAL_CHECK

2. **Non-Product Content Detection (非商品内容识别):**
   - **IMPORTANT**: Check if the description contains non-product content such as:
     - Company success stories, case studies, portfolio pages
     - Customer testimonials or project examples
     - Company information or "about us" content
     - General service descriptions without specific product details
   - If the content is clearly NOT about a specific product → mark as NEEDS_MANUAL_CHECK with reason "Non-product content (e.g., success story, case study)"

**Decision Rules for Product Validity:**
- **PASS**: The product is valid if ALL of the following are true:
  - Content is meaningful and describes a real product (not spam, gibberish, or meaningless)
  - Content is about a specific product (not success stories, case studies, etc.)
  
- **NEEDS_MANUAL_CHECK**: Mark as NEEDS_MANUAL_CHECK if ANY of the following is true:
  - Content is spam, gibberish, or meaningless
  - Content contains non-product information (success stories, case studies, portfolio pages, etc.)

**Category Review:**
- Is the category appropriate for the product described in title and description?
- Does the category match the product type?
- Is the category path logical?
- **Decision Rules:**
  - If category is accurate and appropriate → mark as PASS
  - If category has minor issues (slightly too broad/narrow) → mark as PASS (still acceptable)
  - If category is wrong or significantly inappropriate → mark as NEEDS_MANUAL_CHECK
  - **IMPORTANT**: If category is empty or N/A → mark as NEEDS_MANUAL_CHECK

**Keyword Review:**
- Do the keywords match/describe the product?
- Are keywords relevant to the product?
- **Decision Rules:**
  - If keywords match/describe the product → mark as PASS
  - If some keywords are slightly irrelevant but mostly acceptable → mark as PASS (still acceptable)
  - If keywords are completely irrelevant or don't match the product → mark as NEEDS_MANUAL_CHECK

**Output Format:**
Please provide your review in the following JSON format:
{{
    "product_validity": {{
        "status": "PASS" | "NEEDS_MANUAL_CHECK",
        "reason": "Brief explanation of the review decision"
    }},
    "category_review": {{
        "status": "PASS" | "NEEDS_MANUAL_CHECK",
        "reason": "Brief explanation of the review decision"
    }},
    "keyword_review": {{
        "status": "PASS" | "NEEDS_MANUAL_CHECK",
        "reason": "Brief explanation of the review decision"
    }}
}}

**Status Definitions:**
- "PASS" (通过): The content is acceptable and can be used directly (will be highlighted in green)
- "NEEDS_MANUAL_CHECK" (需要人工复核): The content has significant issues and requires manual review (will be highlighted in red)

**Important:** Please respond ONLY with valid JSON, no additional text or explanations before or after the JSON.

Please provide your review in JSON format only, no additional text."""

        messages = [
            {
                "role": "system",
                "content": "You are a professional product quality auditor. You review product information from online platforms. Since there is no source URL, you must evaluate products based solely on title and description. For product validity, you need to evaluate: non-spam content (meaningful and relevant, not gibberish or placeholder text) and non-product content detection (identify success stories, case studies, portfolio pages, etc.). Product validity has only two statuses: PASS or NEEDS_MANUAL_CHECK. Always respond in valid JSON format."
            },
            {
                "role": "user",
                "content": prompt
            }
        ]
        
        try:
            response = Generation.call(
                model=self.model,
                messages=messages,
                temperature=0.3,
                result_format='message'
            )
            
            if response.status_code == 200:
                # 解析响应
                content = ""
                if hasattr(response, 'output') and response.output is not None:
                    if hasattr(response.output, 'choices') and response.output.choices:
                        if len(response.output.choices) > 0:
                            choice = response.output.choices[0]
                            if hasattr(choice, 'message') and choice.message is not None:
                                if hasattr(choice.message, 'content'):
                                    content = choice.message.content.strip()
                    elif hasattr(response.output, 'text') and response.output.text:
                        content = response.output.text.strip()
                
                if not content:
                    raise ValueError("API响应为空")
                
                # 提取JSON（可能包含markdown代码块）
                if "```json" in content:
                    content = content.split("```json")[1].split("```")[0].strip()
                elif "```" in content:
                    content = content.split("```")[1].split("```")[0].strip()
                
                # 解析JSON
                try:
                    review_result = json.loads(content)
                    
                    # 检查category是否为空或N/A，如果是则直接标记为NEEDS_MANUAL_CHECK
                    if not category_name or not category_name.strip():
                        review_result['category_review'] = {"status": "NEEDS_MANUAL_CHECK", "reason": "Category为空或N/A"}
                    elif category_name.strip().upper() in ['N/A', 'NA', 'NULL', 'NONE']:
                        review_result['category_review'] = {"status": "NEEDS_MANUAL_CHECK", "reason": "Category为N/A"}
                    
                    # 确保包含所有必需的审核项
                    required_reviews = ['product_validity', 'category_review', 'keyword_review']
                    for review_key in required_reviews:
                        if review_key not in review_result:
                            review_result[review_key] = {"status": "NEEDS_MANUAL_CHECK", "reason": f"{review_key}审核结果缺失"}
                    
                    # 确保product_validity只有PASS或NEEDS_MANUAL_CHECK两种状态
                    if 'product_validity' in review_result:
                        status = review_result['product_validity'].get('status', '').upper()
                        if status not in ['PASS', 'NEEDS_MANUAL_CHECK']:
                            # 如果状态不是PASS或NEEDS_MANUAL_CHECK，转换为NEEDS_MANUAL_CHECK
                            review_result['product_validity']['status'] = 'NEEDS_MANUAL_CHECK'
                            review_result['product_validity']['reason'] = f"状态异常，已转换为NEEDS_MANUAL_CHECK。原状态: {status}"
                    
                    return review_result
                except json.JSONDecodeError as e:
                    print(f"JSON解析失败，响应内容前500字符: {content[:500]}")
                    print(f"JSON解析错误: {e}")
                    traceback.print_exc()
                    # 返回默认结果
                    return self._get_default_review("JSON解析失败")
            else:
                error_msg = getattr(response, 'message', '未知错误')
                print(f"API调用失败 (状态码: {response.status_code}): {error_msg}")
                return self._get_default_review(f"API调用失败: {error_msg}")
                
        except Exception as e:
            print(f"审核过程出错: {e}")
            traceback.print_exc()
            return self._get_default_review(f"审核出错: {str(e)}")
    
    def _get_default_review(self, error_msg: str) -> Dict:
        """返回默认审核结果（当出错时）"""
        default_status = "NEEDS_MANUAL_CHECK"
        return {
            "product_validity": {"status": default_status, "reason": error_msg},
            "category_review": {"status": default_status, "reason": error_msg},
            "keyword_review": {"status": default_status, "reason": error_msg}
        }
    
    def _generate_url_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        """根据URL生成汇总统计
        
        Args:
            df: 包含审核结果的DataFrame，必须包含'url'列
            
        Returns:
            按URL分组的汇总统计DataFrame
        """
        if 'url' not in df.columns:
            return pd.DataFrame()
        
        # 过滤掉URL为空的行
        df_with_url = df[df['url'].notna() & (df['url'] != '')].copy()
        
        if len(df_with_url) == 0:
            return pd.DataFrame()
        
        # 按URL分组统计
        summary_data = []
        
        for url, group in df_with_url.groupby('url'):
            total_count = len(group)
            
            # 统计product_validity为PASS的数量
            product_valid_count = len(group[group['product_validity_判定结果'].str.upper() == 'PASS'])
            product_valid_rate = (product_valid_count / total_count * 100) if total_count > 0 else 0
            
            # 统计category为PASS的数量
            category_valid_count = len(group[group['category_判定结果'].str.upper() == 'PASS'])
            category_valid_rate = (category_valid_count / total_count * 100) if total_count > 0 else 0
            
            # 统计keyword为PASS的数量
            keyword_valid_count = len(group[group['keyword_判定结果'].str.upper() == 'PASS'])
            keyword_valid_rate = (keyword_valid_count / total_count * 100) if total_count > 0 else 0
            
            # 统计NEEDS_MANUAL_CHECK的数量
            product_manual_count = len(group[group['product_validity_判定结果'].str.upper() == 'NEEDS_MANUAL_CHECK'])
            category_manual_count = len(group[group['category_判定结果'].str.upper() == 'NEEDS_MANUAL_CHECK'])
            keyword_manual_count = len(group[group['keyword_判定结果'].str.upper() == 'NEEDS_MANUAL_CHECK'])
            
            summary_data.append({
                'url': url,
                '总商品数量': total_count,
                'valid_product数量': product_valid_count,
                'valid_product百分比': f"{product_valid_rate:.2f}%",
                'category_valid数量': category_valid_count,
                'category_valid百分比': f"{category_valid_rate:.2f}%",
                'keyword_valid数量': keyword_valid_count,
                'keyword_valid百分比': f"{keyword_valid_rate:.2f}%",
                'product_需要人工复核数量': product_manual_count,
                'category_需要人工复核数量': category_manual_count,
                'keyword_需要人工复核数量': keyword_manual_count,
            })
        
        summary_df = pd.DataFrame(summary_data)
        
        # 按总商品数量降序排序
        summary_df = summary_df.sort_values('总商品数量', ascending=False).reset_index(drop=True)
        
        return summary_df
    
    def audit_from_csv(self, input_file: str, output_file: str = None):
        """从CSV文件读取并审核商品"""
        
        # 获取当前脚本所在目录（scraper目录）
        current_dir = os.path.dirname(os.path.abspath(__file__))
        report_dir = os.path.join(current_dir, "report_online")
        os.makedirs(report_dir, exist_ok=True)
        
        if output_file is None:
            # 默认保存到report_online文件夹
            base_name = os.path.basename(input_file).replace(".csv", "")
            output_file = os.path.join(report_dir, f"{base_name}_audit_result.xlsx")
        else:
            # 如果指定了输出文件，确保保存到report_online文件夹
            if not os.path.isabs(output_file):
                # 相对路径，保存到report_online文件夹
                if not output_file.startswith("report_online/"):
                    output_file = os.path.join(report_dir, output_file)
                else:
                    output_file = os.path.join(current_dir, output_file)
        
        results = []
        
        # 读取CSV文件
        with open(input_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            total_rows = sum(1 for _ in open(input_file, 'r', encoding='utf-8-sig')) - 1  # 减去header
            
            for idx, row in enumerate(reader, 1):
                # 支持两种格式：online格式和database_merged格式
                # online格式: offer_id, title, description, category_id, category_name, keywords
                # database_merged格式: supplier_id, url, title, cate_info_ai, keyword_ai, description
                
                # 获取ID（优先使用offer_id，如果没有则使用supplier_id）
                offer_id = row.get('offer_id', '') or row.get('supplier_id', '')
                
                title = row.get('title', '')
                description = row.get('description', '')
                
                # 处理category信息
                category_id = row.get('category_id', '')
                category_name = row.get('category_name', '')
                
                # 如果category_id或category_name为空，尝试从cate_info_ai解析
                if not category_id or not category_name:
                    cate_info_ai = row.get('cate_info_ai', '')
                    if cate_info_ai:
                        try:
                            if cate_info_ai.startswith('['):
                                category_list = json.loads(cate_info_ai)
                                if isinstance(category_list, list) and len(category_list) > 0:
                                    category_dict = category_list[0]
                                    if 'catId' in category_dict:
                                        category_id = str(category_dict['catId'])
                                    if 'catPath' in category_dict:
                                        category_name = category_dict['catPath']
                            elif cate_info_ai.startswith('{'):
                                category_dict = json.loads(cate_info_ai)
                                if 'catId' in category_dict:
                                    category_id = str(category_dict['catId'])
                                if 'catPath' in category_dict:
                                    category_name = category_dict['catPath']
                        except:
                            pass  # 如果解析失败，保持为空
                
                # 处理keywords（优先使用keywords，如果没有则使用keyword_ai）
                keywords = row.get('keywords', '') or row.get('keyword_ai', '')
                
                print(f"\n[{idx}/{total_rows}] 审核商品: {title[:50]}...")
                print(f"  ID: {offer_id}")
                
                # 执行审核
                review_result = self.audit_product(
                    offer_id=offer_id,
                    title=title,
                    description=description,
                    category_id=category_id,
                    category_name=category_name,
                    keywords=keywords
                )
                
                # 获取URL（用于后续统计）
                url = row.get('url', '')
                
                # 构建结果行
                result_row = {
                    'id': offer_id,  # 统一使用id作为列名
                    'url': url,  # 添加URL字段
                    'title': title,
                    'description': description[:200] if description else '',  # 限制描述长度
                    'category_id': category_id,
                    'category_name': category_name,
                    'product_validity_判定结果': review_result.get('product_validity', {}).get('status', 'NEEDS_MANUAL_CHECK'),
                    'product_validity_判定原因': review_result.get('product_validity', {}).get('reason', ''),
                    'category_判定结果': review_result.get('category_review', {}).get('status', 'NEEDS_MANUAL_CHECK'),
                    'category_判定原因': review_result.get('category_review', {}).get('reason', ''),
                    'keyword_判定结果': review_result.get('keyword_review', {}).get('status', 'NEEDS_MANUAL_CHECK'),
                    'keyword_判定原因': review_result.get('keyword_review', {}).get('reason', ''),
                }
                
                # 如果存在source_file列，也添加到结果中
                if 'source_file' in row:
                    result_row['source_file'] = row.get('source_file', '')
                
                results.append(result_row)
                
                # 打印简要结果
                print(f"  结果: Product Validity={result_row['product_validity_判定结果']}, "
                      f"Category={result_row['category_判定结果']}, "
                      f"Keyword={result_row['keyword_判定结果']}")
        
        # 保存为Excel并添加颜色格式化
        if results:
            df = pd.DataFrame(results)
            df.to_excel(output_file, index=False, engine='openpyxl')
            
            # 添加颜色格式化
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            
            wb = load_workbook(output_file)
            ws = wb.active
            
            # 定义颜色
            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # 浅绿色
            red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # 红色
            
            # 找到判定结果列的索引
            header_row = 1
            status_columns = {}
            for col_idx, cell in enumerate(ws[header_row], 1):
                if cell.value and '判定结果' in str(cell.value):
                    status_columns[col_idx] = cell.value
            
            # 为判定结果单元格着色
            for row_idx in range(2, len(results) + 2):  # 从第2行开始（跳过header）
                for col_idx, col_name in status_columns.items():
                    cell = ws.cell(row=row_idx, column=col_idx)
                    status = str(cell.value).upper()
                    
                    if status == "PASS":
                        cell.fill = green_fill
                    elif status == "NEEDS_MANUAL_CHECK":
                        cell.fill = red_fill
            
            # 保存格式化后的Excel
            wb.save(output_file)
            
            print(f"\n✅ 审核完成！结果已保存到: {output_file}")
            print(f"共审核 {len(results)} 个商品")
            
            # 打印统计信息
            print("\n📊 统计信息:")
            for col in ['product_validity_判定结果', 'category_判定结果', 'keyword_判定结果']:
                if col in df.columns:
                    status_counts = df[col].value_counts()
                    print(f"\n{col}:")
                    for status, count in status_counts.items():
                        percentage = (count / len(df)) * 100
                        print(f"  {status}: {count} ({percentage:.1f}%)")
            
            # 如果存在URL列，生成按URL分组的汇总统计
            if 'url' in df.columns and df['url'].notna().any():
                print("\n📈 正在生成URL汇总统计...")
                url_summary = self._generate_url_summary(df)
                
                if len(url_summary) > 0:
                    # 保存URL汇总统计
                    if output_file.endswith('.xlsx'):
                        summary_output_file = output_file.replace('.xlsx', '_url_summary.xlsx')
                    else:
                        summary_output_file = output_file + '_url_summary.xlsx'
                    
                    url_summary.to_excel(summary_output_file, index=False, engine='openpyxl')
                    
                    # 格式化汇总文件
                    from openpyxl import load_workbook
                    from openpyxl.styles import PatternFill, Font, Alignment
                    
                    wb_summary = load_workbook(summary_output_file)
                    ws_summary = wb_summary.active
                    
                    # 设置标题行格式
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    for cell in ws_summary[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 调整列宽
                    for column in ws_summary.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws_summary.column_dimensions[column_letter].width = adjusted_width
                    
                    wb_summary.save(summary_output_file)
                    print(f"✅ URL汇总统计已保存到: {summary_output_file}")
                    print(f"   共统计 {len(url_summary)} 个URL")
                else:
                    print("⚠️ 没有有效的URL数据可统计")
        else:
            print("⚠️ 没有审核结果可保存")


def main():
    """主函数"""
    import argparse
    
    parser = argparse.ArgumentParser(description='在线商品审计工具（无URL版本）')
    parser.add_argument('input_file', help='输入CSV文件路径（如: online/old_url_scrap_data_output.csv）')
    parser.add_argument('-o', '--output', help='输出Excel文件路径（可选，默认为输入文件名_audit_result.xlsx）')
    parser.add_argument('--api-key', help='Qwen/DashScope API Key（可选，也可通过环境变量QWEN_API_KEY或DASHSCOPE_API_KEY设置）')
    parser.add_argument('--model', default='qwen-plus', help='使用的模型名称（默认: qwen-plus，可选: qwen-turbo, qwen-max等）')
    
    args = parser.parse_args()
    
    # 检查API Key
    api_key = args.api_key or os.getenv("QWEN_API_KEY") or os.getenv("DASHSCOPE_API_KEY")
    if not api_key:
        current_dir = os.path.dirname(os.path.abspath(__file__))
        root_dir = os.path.dirname(current_dir)
        env_path = os.path.join(root_dir, '.env')
        print(f"错误: 请设置QWEN_API_KEY或DASHSCOPE_API_KEY环境变量或使用--api-key参数")
        print(f"提示: 环境变量文件位置: {env_path}")
        return
    
    # 创建审计员并执行审核
    auditor = ProductAuditorOnline(api_key=api_key, model=args.model)
    auditor.audit_from_csv(args.input_file, args.output)


if __name__ == "__main__":
    main()

