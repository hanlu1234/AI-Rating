"""
商品审计工具
使用AI审核从爬虫获取的商品信息（title, description, image, category, keyword）
"""

import csv
import json
import os
import sys
import traceback
from typing import Dict, List, Optional
import dashscope
from dashscope import Generation
from dotenv import load_dotenv
import pandas as pd
from datetime import datetime

# 加载环境变量（从项目根目录加载）
# 获取项目根目录（ai rating目录）
current_dir = os.path.dirname(os.path.abspath(__file__))
root_dir = os.path.dirname(current_dir)  # 上一级目录就是ai rating
env_path = os.path.join(root_dir, '.env')
load_dotenv(env_path)


class ProductAuditor:
    """商品审计员"""
    
    def __init__(self, api_key: str = None, model: str = "qwen-plus"):
        """初始化审计员
        
        Args:
            api_key: DashScope API Key
            model: 使用的模型名称，默认为 qwen-plus
        """
        self.api_key = api_key or os.getenv("QWEN_API_KEY") or os.getenv("DASHSCOPE_API_KEY")
        if not self.api_key:
            # 获取env_path用于错误提示
            current_dir = os.path.dirname(os.path.abspath(__file__))
            root_dir = os.path.dirname(current_dir)
            env_path = os.path.join(root_dir, '.env')
            raise ValueError(f"API Key未设置，请通过参数传入或设置环境变量QWEN_API_KEY或DASHSCOPE_API_KEY。环境变量文件位置: {env_path}")
        dashscope.api_key = self.api_key
        self.model = model
    
    def audit_product(self, url: str, title: str, description: str, 
                     main_image: str, image_list: str, 
                     category: str, keyword: str) -> Dict:
        """审核单个商品
        
        Args:
            url: 商品源URL
            title: 商品标题
            description: 商品描述
            main_image: 主图URL
            image_list: 其他图片URL列表（JSON字符串或逗号分隔）
            category: AI预测的category
            keyword: AI预测的keyword（JSON字符串）
            
        Returns:
            包含各项审核结果的字典
        """
        
        # 解析image_list
        image_urls = []
        if image_list:
            try:
                if image_list.startswith('[') or image_list.startswith('{'):
                    parsed = json.loads(image_list)
                    if isinstance(parsed, list):
                        image_urls = parsed
                    elif isinstance(parsed, dict):
                        image_urls = list(parsed.values())
                else:
                    # 逗号分隔的字符串
                    image_urls = [img.strip() for img in image_list.split(',') if img.strip()]
            except:
                image_urls = [img.strip() for img in image_list.split(',') if img.strip()] if image_list else []
        
        # 解析keyword
        keyword_text = ""
        if keyword:
            try:
                if keyword.startswith('{'):
                    keyword_dict = json.loads(keyword)
                    if 'keywords_english' in keyword_dict:
                        keyword_text = ', '.join(keyword_dict['keywords_english'].values())
                    elif 'keywords' in keyword_dict:
                        keyword_text = ', '.join(keyword_dict['keywords'].values())
                    else:
                        keyword_text = str(keyword_dict)
                else:
                    keyword_text = keyword
            except:
                keyword_text = keyword
        
        # 解析category
        category_text = ""
        if category:
            try:
                if category.startswith('['):
                    category_list = json.loads(category)
                    if isinstance(category_list, list) and len(category_list) > 0:
                        if 'catPath' in category_list[0]:
                            category_text = category_list[0]['catPath']
                        else:
                            category_text = str(category_list[0])
                elif category.startswith('{'):
                    category_dict = json.loads(category)
                    if 'catPath' in category_dict:
                        category_text = category_dict['catPath']
                    else:
                        category_text = str(category_dict)
                else:
                    category_text = category
            except:
                category_text = category
        
        # 检查category是否为空或N/A，如果是则直接标记为NEEDS_MANUAL_CHECK
        category_is_empty_or_na = False
        if not category_text:
            category_is_empty_or_na = True
        else:
            category_text_trimmed = category_text.strip()
            if not category_text_trimmed:
                category_is_empty_or_na = True
            elif category_text_trimmed.upper() in ['N/A', 'NA', 'NULL', 'NONE']:
                category_is_empty_or_na = True
        
        # 构建审核prompt
        prompt = f"""You are an AI product auditor. Please review the following product information scraped from a website and imported to a new platform.

**Product Source URL:** {url}

**Product Information to Review:**

0. **Product URL** (scraped URL):
{url if url else "N/A"}

1. **Title** (scraped from source website):
{title if title else "N/A"}

2. **Description** (scraped from source website):
{description if description else "N/A"}

3. **AI Predicted Category:**
{category_text if category_text else "N/A"}

4. **AI Predicted Keywords:**
{keyword_text if keyword_text else "N/A"}

**Review Criteria:**

For each aspect (URL, Title, Description, Category, Keyword), please evaluate:
Note: Image review is currently skipped.

**URL Review:**
- Is the URL a valid product page URL?
- Does the URL structure indicate it's a product page (not a category page, homepage, or other non-product page)?
- **IMPORTANT**: Check if the URL is a company success case/portfolio page (e.g., contains "case", "portfolio", "success", "project", "client", "example", "story", etc.). If it's a success case page, it is NOT a product URL and should be marked as NEEDS_MANUAL_CHECK.
- **IMPORTANT**: Check if the URL is a category page or multi-product listing page (e.g., contains "category", "catalog", "products", "list", "collection", "browse", "shop", "all", or shows multiple products). If it's a category/multi-product page, mark as NEEDS_REVIEW and specify "Category page" or "Multi-product page" in the reason (this will be highlighted in yellow).
- Are there any other signs that this is not a product URL (e.g., contains "search", "home", "about", "contact", etc.)?
- **Decision Rules:**
  - If the URL is clearly a valid single product page and is NOT a success case page → mark as PASS
  - If the URL is a success case page → mark as NEEDS_MANUAL_CHECK
  - If the URL is a category page or multi-product listing page → mark as NEEDS_REVIEW with reason "Category page" or "Multi-product page" (this will be highlighted in yellow)
  - If you are uncertain or unclear whether it's a product URL → mark as NEEDS_REVIEW (this will be highlighted in yellow)

**Title Review:**
Since the title is directly scraped from the source website without AI processing, please check:
- Does the scraped title match the product name/identifier in the source URL?
- Is the title consistent with what would be expected from the URL structure?
- If the URL contains product identifiers or names, do they match the scraped title?
- If there's a mismatch or inconsistency, mark as NEEDS_REVIEW or NEEDS_MANUAL_CHECK accordingly.

**Description Review:**
Since the description is directly scraped from the source website without AI processing, please check:
- Does the scraped description match/relate to the product name in the URL?
- Is the description consistent with the product title?
- Does the description content align with what would be expected for this product based on the URL?
- If there's a mismatch or inconsistency, mark as NEEDS_REVIEW or NEEDS_MANUAL_CHECK accordingly.

**Category Review:**
- Is the AI-predicted category accurate and appropriate?
- Does it match the product type?
- Is the category path logical?
- Are there any issues (wrong category, too broad, too narrow)?

**Keyword Review:**
Since keywords are limited to a maximum of 3, the review criteria is simplified:
- Do the keywords match/describe the product? If yes, then PASS.
- If keywords are irrelevant or don't match the product description, then mark as NEEDS_REVIEW or NEEDS_MANUAL_CHECK based on severity.

**Output Format:**
Please provide your review in the following JSON format:
{{
    "url_review": {{
        "status": "PASS" | "NEEDS_REVIEW" | "NEEDS_MANUAL_CHECK",
        "reason": "Brief explanation of the review decision"
    }},
    "title_review": {{
        "status": "PASS" | "NEEDS_REVIEW" | "NEEDS_MANUAL_CHECK",
        "reason": "Brief explanation of the review decision"
    }},
    "description_review": {{
        "status": "PASS" | "NEEDS_REVIEW" | "NEEDS_MANUAL_CHECK",
        "reason": "Brief explanation of the review decision"
    }},
    "category_review": {{
        "status": "PASS" | "NEEDS_REVIEW" | "NEEDS_MANUAL_CHECK",
        "reason": "Brief explanation of the review decision"
    }},
    "keyword_review": {{
        "status": "PASS" | "NEEDS_REVIEW" | "NEEDS_MANUAL_CHECK",
        "reason": "Brief explanation of the review decision"
    }}
}}

**Status Definitions:**
- "PASS" (通过): The content is acceptable and can be used directly (will be highlighted in green)
  - For Keywords: If keywords match/describe the product, mark as PASS
  - For URL: If the URL is clearly a valid product page
- "NEEDS_REVIEW" (需要抽查): The content has minor issues, is uncertain, or needs spot-checking (will be highlighted in yellow)
  - For Keywords: If some keywords are slightly irrelevant but mostly acceptable
  - For URL: 
    - If you are uncertain whether the URL is a valid product page
    - If the URL is a category page or multi-product listing page (mark as NEEDS_REVIEW with reason "Category page" or "Multi-product page")
  - For Title/Description: If there's slight inconsistency but mostly acceptable
- "NEEDS_MANUAL_CHECK" (需要人工复核): The content has significant issues and requires manual review (will be highlighted in red)
  - For Keywords: If keywords are completely irrelevant or don't match the product at all
  - For URL: If the URL is clearly NOT a product page (e.g., success case page, category page, etc.)
  - For Title/Description: If there's significant inconsistency or mismatch

**Important:** Please respond ONLY with valid JSON, no additional text or explanations before or after the JSON.

Please provide your review in JSON format only, no additional text."""

        messages = [
            {
                "role": "system",
                "content": "You are a professional product quality auditor. You review scraped product information for consistency with the source website. The title and description are directly scraped from the source website without AI processing, so you should focus on checking consistency between the URL, title, and description. Always respond in valid JSON format."
            },
            {
                "role": "user",
                "content": prompt
            }
        ]
        
        try:
            response = Generation.call(
                model=self.model,
                messages=messages,
                temperature=0.3,
                result_format='message'
            )
            
            if response.status_code == 200:
                # 解析响应
                content = ""
                if hasattr(response, 'output') and response.output is not None:
                    if hasattr(response.output, 'choices') and response.output.choices:
                        if len(response.output.choices) > 0:
                            choice = response.output.choices[0]
                            if hasattr(choice, 'message') and choice.message is not None:
                                if hasattr(choice.message, 'content'):
                                    content = choice.message.content.strip()
                    elif hasattr(response.output, 'text') and response.output.text:
                        content = response.output.text.strip()
                
                if not content:
                    raise ValueError("API响应为空")
                
                # 提取JSON（可能包含markdown代码块）
                if "```json" in content:
                    content = content.split("```json")[1].split("```")[0].strip()
                elif "```" in content:
                    content = content.split("```")[1].split("```")[0].strip()
                
                # 解析JSON
                try:
                    review_result = json.loads(content)
                    # 确保不包含image_review（已跳过）
                    if 'image_review' in review_result:
                        del review_result['image_review']
                    # 确保包含url_review
                    if 'url_review' not in review_result:
                        review_result['url_review'] = {"status": "NEEDS_MANUAL_CHECK", "reason": "URL审核结果缺失"}
                    # 如果category为空或N/A，直接标记为NEEDS_MANUAL_CHECK
                    if category_is_empty_or_na:
                        review_result['category_review'] = {"status": "NEEDS_MANUAL_CHECK", "reason": "Category为空或N/A"}
                    return review_result
                except json.JSONDecodeError as e:
                    print(f"JSON解析失败，响应内容前500字符: {content[:500]}")
                    print(f"JSON解析错误: {e}")
                    traceback.print_exc()
                    # 返回默认结果
                    return self._get_default_review("JSON解析失败")
            else:
                error_msg = getattr(response, 'message', '未知错误')
                print(f"API调用失败 (状态码: {response.status_code}): {error_msg}")
                return self._get_default_review(f"API调用失败: {error_msg}")
                
        except Exception as e:
            print(f"审核过程出错: {e}")
            traceback.print_exc()
            return self._get_default_review(f"审核出错: {str(e)}")
    
    def _get_default_review(self, error_msg: str) -> Dict:
        """返回默认审核结果（当出错时）"""
        default_status = "NEEDS_MANUAL_CHECK"
        return {
            "url_review": {"status": default_status, "reason": error_msg},
            "title_review": {"status": default_status, "reason": error_msg},
            "description_review": {"status": default_status, "reason": error_msg},
            "image_review": {"status": "SKIP", "reason": "Image审核已跳过"},
            "category_review": {"status": default_status, "reason": error_msg},
            "keyword_review": {"status": default_status, "reason": error_msg}
        }
    
    def audit_from_csv(self, input_file: str, output_file: str = None):
        """从CSV文件读取并审核商品"""
        
        # 获取当前脚本所在目录（scraper目录）
        current_dir = os.path.dirname(os.path.abspath(__file__))
        report_dir = os.path.join(current_dir, "report")
        os.makedirs(report_dir, exist_ok=True)
        
        if output_file is None:
            # 默认保存到report文件夹
            base_name = os.path.basename(input_file).replace(".csv", "")
            output_file = os.path.join(report_dir, f"{base_name}_audit_result.xlsx")
        else:
            # 如果指定了输出文件，确保保存到report文件夹
            if not os.path.isabs(output_file):
                # 相对路径，保存到report文件夹
                if not output_file.startswith("report/"):
                    output_file = os.path.join(report_dir, output_file)
                else:
                    output_file = os.path.join(current_dir, output_file)
        
        results = []
        
        # 读取CSV文件
        with open(input_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            total_rows = sum(1 for _ in open(input_file, 'r', encoding='utf-8-sig')) - 1  # 减去header
            
            for idx, row in enumerate(reader, 1):
                url = row.get('url', '')
                title = row.get('title', '')
                description = row.get('description', '')
                main_image = row.get('product_main_image', '')
                image_list = row.get('product_image_list', '')
                category = row.get('cate_info_ai', '')
                keyword = row.get('keyword_ai', '')
                
                print(f"\n[{idx}/{total_rows}] 审核商品: {title[:50]}...")
                print(f"  URL: {url}")
                
                # 执行审核
                review_result = self.audit_product(
                    url=url,
                    title=title,
                    description=description,
                    main_image=main_image,
                    image_list=image_list,
                    category=category,
                    keyword=keyword
                )
                
                # 构建结果行（按照用户要求：包含title和各类判定结果）
                # 注意：image审核已跳过
                result_row = {
                    'id': row.get('id', ''),
                    'url': url,
                    'title': title,
                    'url_判定结果': review_result.get('url_review', {}).get('status', 'NEEDS_MANUAL_CHECK'),
                    'url_判定原因': review_result.get('url_review', {}).get('reason', ''),
                    'title_判定结果': review_result.get('title_review', {}).get('status', 'NEEDS_MANUAL_CHECK'),
                    'title_判定原因': review_result.get('title_review', {}).get('reason', ''),
                    'description_判定结果': review_result.get('description_review', {}).get('status', 'NEEDS_MANUAL_CHECK'),
                    'description_判定原因': review_result.get('description_review', {}).get('reason', ''),
                    'category_判定结果': review_result.get('category_review', {}).get('status', 'NEEDS_MANUAL_CHECK'),
                    'category_判定原因': review_result.get('category_review', {}).get('reason', ''),
                    'keyword_判定结果': review_result.get('keyword_review', {}).get('status', 'NEEDS_MANUAL_CHECK'),
                    'keyword_判定原因': review_result.get('keyword_review', {}).get('reason', ''),
                }
                
                results.append(result_row)
                
                # 打印简要结果
                print(f"  结果: URL={result_row['url_判定结果']}, Title={result_row['title_判定结果']}, Description={result_row['description_判定结果']}, "
                      f"Category={result_row['category_判定结果']}, Keyword={result_row['keyword_判定结果']}")
        
        # 保存为Excel并添加颜色格式化
        if results:
            df = pd.DataFrame(results)
            df.to_excel(output_file, index=False, engine='openpyxl')
            
            # 添加颜色格式化
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            
            wb = load_workbook(output_file)
            ws = wb.active
            
            # 定义颜色
            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # 浅绿色
            yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # 黄色
            red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # 红色
            
            # 找到判定结果列的索引
            header_row = 1
            status_columns = {}
            for col_idx, cell in enumerate(ws[header_row], 1):
                if cell.value and '判定结果' in str(cell.value):
                    status_columns[col_idx] = cell.value
            
            # 为判定结果单元格着色
            for row_idx in range(2, len(results) + 2):  # 从第2行开始（跳过header）
                for col_idx, col_name in status_columns.items():
                    cell = ws.cell(row=row_idx, column=col_idx)
                    status = str(cell.value).upper()
                    
                    if status == "PASS":
                        cell.fill = green_fill
                    elif status == "NEEDS_REVIEW":
                        cell.fill = yellow_fill
                    elif status == "NEEDS_MANUAL_CHECK":
                        cell.fill = red_fill
            
            # 保存格式化后的Excel
            wb.save(output_file)
            
            print(f"\n✅ 审核完成！结果已保存到: {output_file}")
            print(f"共审核 {len(results)} 个商品")
            
            # 打印统计信息
            print("\n📊 统计信息:")
            for col in ['title_判定结果', 'description_判定结果', 'category_判定结果', 'keyword_判定结果']:
                if col in df.columns:
                    status_counts = df[col].value_counts()
                    print(f"\n{col}:")
                    for status, count in status_counts.items():
                        percentage = (count / len(df)) * 100
                        print(f"  {status}: {count} ({percentage:.1f}%)")
        else:
            print("⚠️ 没有审核结果可保存")


def main():
    """主函数"""
    import argparse
    
    parser = argparse.ArgumentParser(description='商品审计工具')
    parser.add_argument('input_file', help='输入CSV文件路径（如: database/dema.csv）')
    parser.add_argument('-o', '--output', help='输出Excel文件路径（可选，默认为输入文件名_audit_result.xlsx）')
    parser.add_argument('--api-key', help='Qwen/DashScope API Key（可选，也可通过环境变量QWEN_API_KEY或DASHSCOPE_API_KEY设置）')
    parser.add_argument('--model', default='qwen-plus', help='使用的模型名称（默认: qwen-plus，可选: qwen-turbo, qwen-max等）')
    
    args = parser.parse_args()
    
    # 检查API Key
    api_key = args.api_key or os.getenv("QWEN_API_KEY") or os.getenv("DASHSCOPE_API_KEY")
    if not api_key:
        current_dir = os.path.dirname(os.path.abspath(__file__))
        root_dir = os.path.dirname(current_dir)
        env_path = os.path.join(root_dir, '.env')
        print(f"错误: 请设置QWEN_API_KEY或DASHSCOPE_API_KEY环境变量或使用--api-key参数")
        print(f"提示: 环境变量文件位置: {env_path}")
        return
    
    # 创建审计员并执行审核
    auditor = ProductAuditor(api_key=api_key, model=args.model)
    auditor.audit_from_csv(args.input_file, args.output)


if __name__ == "__main__":
    main()

