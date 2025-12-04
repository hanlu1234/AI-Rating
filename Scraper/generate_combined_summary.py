"""
Generate a combined summary report from all audit result files
"""

import pandas as pd
import os
import glob
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def generate_combined_summary(report_dir="report", output_file=None):
    """Generate a combined summary report from all audit result files"""
    
    if not os.path.exists(report_dir):
        print(f"Error: Report directory not found: {report_dir}")
        return
    
    # Find all audit result files
    audit_files = glob.glob(os.path.join(report_dir, "*_audit_result.xlsx"))
    
    if not audit_files:
        print(f"No audit result files found in {report_dir}")
        return
    
    print(f"Found {len(audit_files)} audit result file(s)")
    print("=" * 60)
    
    # Combine all audit results
    all_dataframes = []
    file_names = []
    
    for audit_file in sorted(audit_files):
        file_name = os.path.basename(audit_file).replace("_audit_result.xlsx", "")
        print(f"Loading: {file_name}")
        try:
            df = pd.read_excel(audit_file, engine='openpyxl')
            df['Source File'] = file_name  # Add source file column
            all_dataframes.append(df)
            file_names.append(file_name)
        except Exception as e:
            print(f"  ❌ Error loading {audit_file}: {e}")
    
    if not all_dataframes:
        print("No valid audit files found")
        return
    
    # Combine all dataframes
    combined_df = pd.concat(all_dataframes, ignore_index=True)
    total_products = len(combined_df)
    
    print(f"\n✅ Combined {len(all_dataframes)} file(s), total {total_products} products")
    
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(report_dir, f"combined_summary_report_{timestamp}.xlsx")
    
    # Create summary statistics
    summary_data = []
    
    # Get all status columns
    status_columns = [col for col in combined_df.columns if '判定结果' in col]
    
    for col in status_columns:
        aspect = col.replace('_判定结果', '').replace('_', ' ').title()
        status_counts = combined_df[col].value_counts()
        
        pass_count = status_counts.get('PASS', 0)
        review_count = status_counts.get('NEEDS_REVIEW', 0)
        manual_count = status_counts.get('NEEDS_MANUAL_CHECK', 0)
        
        pass_rate = (pass_count / total_products * 100) if total_products > 0 else 0
        review_rate = (review_count / total_products * 100) if total_products > 0 else 0
        manual_rate = (manual_count / total_products * 100) if total_products > 0 else 0
        
        summary_data.append({
            'Aspect': aspect,
            'Total': total_products,
            'PASS': pass_count,
            'PASS Rate (%)': f"{pass_rate:.1f}%",
            'Needs Review': review_count,
            'Review Rate (%)': f"{review_rate:.1f}%",
            'Needs Manual Check': manual_count,
            'Manual Rate (%)': f"{manual_rate:.1f}%"
        })
    
    # Create detailed issue analysis
    issue_analysis = []
    
    for col in status_columns:
        aspect = col.replace('_判定结果', '').replace('_', ' ').title()
        reason_col = col.replace('_判定结果', '_判定原因')
        
        if reason_col in combined_df.columns:
            # Get products that need attention
            needs_attention = combined_df[combined_df[col].isin(['NEEDS_REVIEW', 'NEEDS_MANUAL_CHECK'])]
            
            if len(needs_attention) > 0:
                # Count by status
                review_items = needs_attention[needs_attention[col] == 'NEEDS_REVIEW']
                manual_items = needs_attention[needs_attention[col] == 'NEEDS_MANUAL_CHECK']
                
                # Get common reasons
                if len(review_items) > 0:
                    common_reasons = review_items[reason_col].value_counts().head(10)
                    for reason, count in common_reasons.items():
                        issue_analysis.append({
                            'Aspect': aspect,
                            'Status': 'NEEDS_REVIEW',
                            'Issue': str(reason)[:150] if pd.notna(reason) else 'N/A',
                            'Count': count,
                            'Percentage': f"{count / total_products * 100:.1f}%"
                        })
                
                if len(manual_items) > 0:
                    common_reasons = manual_items[reason_col].value_counts().head(10)
                    for reason, count in common_reasons.items():
                        issue_analysis.append({
                            'Aspect': aspect,
                            'Status': 'NEEDS_MANUAL_CHECK',
                            'Issue': str(reason)[:150] if pd.notna(reason) else 'N/A',
                            'Count': count,
                            'Percentage': f"{count / total_products * 100:.1f}%"
                        })
    
    # Statistics by source file
    file_stats = []
    for file_name in file_names:
        file_df = combined_df[combined_df['Source File'] == file_name]
        file_total = len(file_df)
        
        file_stat = {'Source File': file_name, 'Total Products': file_total}
        
        for col in status_columns:
            aspect = col.replace('_判定结果', '').replace('_', ' ').title()
            status_counts = file_df[col].value_counts()
            pass_count = status_counts.get('PASS', 0)
            pass_rate = (pass_count / file_total * 100) if file_total > 0 else 0
            file_stat[f'{aspect} Pass Rate'] = f"{pass_rate:.1f}%"
        
        file_stats.append(file_stat)
    
    # Create Excel workbook with multiple sheets
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Summary Statistics Sheet
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary Statistics', index=False)
        
        # File Statistics Sheet
        file_stats_df = pd.DataFrame(file_stats)
        file_stats_df.to_excel(writer, sheet_name='File Statistics', index=False)
        
        # Issue Analysis Sheet
        if issue_analysis:
            issue_df = pd.DataFrame(issue_analysis)
            issue_df.to_excel(writer, sheet_name='Issue Analysis', index=False)
        else:
            pd.DataFrame({'Message': ['No issues found']}).to_excel(writer, sheet_name='Issue Analysis', index=False)
        
        # Problem Items Sheet (all items that need attention)
        problem_items = combined_df[combined_df[status_columns].isin(['NEEDS_REVIEW', 'NEEDS_MANUAL_CHECK']).any(axis=1)]
        if len(problem_items) > 0:
            # Select key columns
            key_columns = ['Source File', 'id', 'url', 'title'] + status_columns + [col.replace('_判定结果', '_判定原因') for col in status_columns if col.replace('_判定结果', '_判定原因') in combined_df.columns]
            available_columns = [col for col in key_columns if col in problem_items.columns]
            problem_items[available_columns].to_excel(writer, sheet_name='Problem Items', index=False)
        else:
            pd.DataFrame({'Message': ['No problematic items found']}).to_excel(writer, sheet_name='Problem Items', index=False)
        
        # All Combined Data Sheet
        key_columns = ['Source File', 'id', 'url', 'title'] + status_columns + [col.replace('_判定结果', '_判定原因') for col in status_columns if col.replace('_判定结果', '_判定原因') in combined_df.columns]
        available_columns = [col for col in key_columns if col in combined_df.columns]
        combined_df[available_columns].to_excel(writer, sheet_name='All Data', index=False)
    
    # Format the Excel file
    wb = load_workbook(output_file)
    
    # Format all sheets
    format_summary_sheet(wb['Summary Statistics'])
    if 'File Statistics' in wb.sheetnames:
        format_file_stats_sheet(wb['File Statistics'])
    if 'Issue Analysis' in wb.sheetnames:
        format_issue_sheet(wb['Issue Analysis'])
    if 'Problem Items' in wb.sheetnames:
        format_problem_sheet(wb['Problem Items'], status_columns)
    if 'All Data' in wb.sheetnames:
        format_all_data_sheet(wb['All Data'], status_columns)
    
    wb.save(output_file)
    
    print(f"\n✅ Combined summary report generated: {output_file}")
    print(f"📊 Total products: {total_products}")
    print(f"📁 Report contains {len(wb.sheetnames)} sheets:")
    for sheet_name in wb.sheetnames:
        print(f"   - {sheet_name}")
    
    return output_file


def format_summary_sheet(ws):
    """Format the summary statistics sheet"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws.column_dimensions['A'].width = 20
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    for row in range(2, ws.max_row + 1):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if col == 4:  # PASS Rate column
                try:
                    rate = float(cell.value.replace('%', ''))
                    if rate >= 80:
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif rate >= 60:
                        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                except:
                    pass


def format_file_stats_sheet(ws):
    """Format the file statistics sheet"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    for col in range(3, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')


def format_issue_sheet(ws):
    """Format the issue analysis sheet"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    
    for row in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row, column=2)
        if status_cell.value == 'NEEDS_REVIEW':
            status_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        elif status_cell.value == 'NEEDS_MANUAL_CHECK':
            status_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col != 3:  # Don't center the issue description
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


def format_problem_sheet(ws, status_columns):
    """Format the problem items sheet"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws.column_dimensions['A'].width = 20  # Source File
    ws.column_dimensions['B'].width = 10  # ID
    ws.column_dimensions['C'].width = 50  # URL
    ws.column_dimensions['D'].width = 40  # Title
    
    for row in range(2, ws.max_row + 1):
        for col_idx, cell in enumerate(ws[row], 1):
            header_cell = ws.cell(row=1, column=col_idx)
            if cell.value and header_cell.value and '判定结果' in str(header_cell.value):
                status = str(cell.value).upper()
                if status == "PASS":
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif status == "NEEDS_REVIEW":
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                elif status == "NEEDS_MANUAL_CHECK":
                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            
            cell.border = border
            if col_idx <= 4:  # Source File, ID, URL, Title
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')


def format_all_data_sheet(ws, status_columns):
    """Format the all data sheet"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws.column_dimensions['A'].width = 20  # Source File
    ws.column_dimensions['B'].width = 10  # ID
    ws.column_dimensions['C'].width = 50  # URL
    ws.column_dimensions['D'].width = 40  # Title
    
    for row in range(2, ws.max_row + 1):
        for col_idx, cell in enumerate(ws[row], 1):
            header_cell = ws.cell(row=1, column=col_idx)
            if cell.value and header_cell.value and '判定结果' in str(header_cell.value):
                status = str(cell.value).upper()
                if status == "PASS":
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif status == "NEEDS_REVIEW":
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                elif status == "NEEDS_MANUAL_CHECK":
                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            
            cell.border = border
            if col_idx <= 4:  # Source File, ID, URL, Title
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')


def main():
    """Main function"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Generate combined summary report from all audit result files')
    parser.add_argument('-d', '--dir', default='report', help='Report directory (default: report)')
    parser.add_argument('-o', '--output', help='Output file path (optional)')
    
    args = parser.parse_args()
    
    generate_combined_summary(args.dir, args.output)


if __name__ == "__main__":
    main()


